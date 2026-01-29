#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

MAX_EXCEL_ROWS = 1_048_576     # per sheet, incl header
MAX_EXCEL_COLS = 16_384        # per sheet


def detect_encoding(path: Path) -> str:
    """Try utf-8, fall back to latin-1."""
    try:
        with path.open("r", encoding="utf-8") as f:
            f.readline()
        return "utf-8"
    except UnicodeDecodeError:
        return "latin-1"


def read_header_cols(path: Path, encoding: str) -> list[str]:
    with path.open("r", encoding=encoding, newline="") as f:
        header = f.readline().rstrip("\r\n")
    return header.split("\t") if header else []


def safe_sheet_name(base: str, idx: int) -> str:
    # Excel sheet name max length is 31, no : \ / ? * [ ]
    cleaned = "".join(ch for ch in base if ch not in r':\/?*[]')
    name = f"{cleaned[:20]}_{idx:03d}"  # keep it short
    return name[:31]


def write_df_rows(ws, df: pd.DataFrame) -> None:
    # write_only + ws.append wants plain python values
    for row in df.itertuples(index=False, name=None):
        ws.append([("" if v is None else v) for v in row])


def convert_tsv_split(
    tsv_path: Path,
    out_dir: Path,
    mode: str = "sheets",
    chunksize: int = 50_000,
) -> tuple[str, str, int, str]:
    """
    Returns: (status, reason, parts, outputs_string)
    """
    encoding = detect_encoding(tsv_path)
    header_cols = read_header_cols(tsv_path, encoding=encoding)
    cols = len(header_cols)

    if cols == 0:
        return ("SKIPPED", "Empty or missing header", 0, "")

    if cols > MAX_EXCEL_COLS:
        return ("SKIPPED", f"Too many columns for Excel (max {MAX_EXCEL_COLS:,})", 0, "")

    max_data_rows_per_sheet = MAX_EXCEL_ROWS - 1  # reserve 1 row for header

    reader = pd.read_csv(
        tsv_path,
        sep="\t",
        dtype=str,
        encoding=encoding,
        keep_default_na=False,
        chunksize=chunksize,
    )

    outputs: list[str] = []

    if mode == "sheets":
        wb = Workbook(write_only=True)
        sheet_idx = 1
        ws = wb.create_sheet(title=safe_sheet_name(tsv_path.stem, sheet_idx))
        ws.append(header_cols)
        rows_in_sheet = 1  # header already written

        total_rows_written = 0

        for chunk in reader:
            # chunk columns from pandas should match header; keep them in same order
            chunk = chunk.reindex(columns=header_cols)

            start = 0
            n = len(chunk)
            while start < n:
                remaining_capacity = max_data_rows_per_sheet - (rows_in_sheet - 1)
                if remaining_capacity <= 0:
                    sheet_idx += 1
                    ws = wb.create_sheet(title=safe_sheet_name(tsv_path.stem, sheet_idx))
                    ws.append(header_cols)
                    rows_in_sheet = 1
                    remaining_capacity = max_data_rows_per_sheet

                take = min(remaining_capacity, n - start)
                sub = chunk.iloc[start : start + take]
                write_df_rows(ws, sub)

                rows_in_sheet += take
                total_rows_written += take
                start += take

        xlsx_path = out_dir / f"{tsv_path.stem}.xlsx"
        wb.save(xlsx_path)
        outputs.append(str(xlsx_path))

        if sheet_idx > 1:
            return ("SPLIT", f"Split into {sheet_idx} sheet(s)", sheet_idx, ";".join(outputs))
        return ("OK", "", 1, ";".join(outputs))

    elif mode == "files":
        part_idx = 1
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Sheet1")
        ws.append(header_cols)
        rows_in_part = 1  # header
        parts_created = 1

        def save_part(workbook: Workbook, idx: int) -> Path:
            xlsx_path = out_dir / f"{tsv_path.stem}_part{idx:03d}.xlsx"
            workbook.save(xlsx_path)
            return xlsx_path

        for chunk in reader:
            chunk = chunk.reindex(columns=header_cols)

            start = 0
            n = len(chunk)
            while start < n:
                remaining_capacity = max_data_rows_per_sheet - (rows_in_part - 1)
                if remaining_capacity <= 0:
                    # save current part, start a new workbook
                    outputs.append(str(save_part(wb, part_idx)))
                    part_idx += 1
                    parts_created += 1

                    wb = Workbook(write_only=True)
                    ws = wb.create_sheet(title="Sheet1")
                    ws.append(header_cols)
                    rows_in_part = 1
                    remaining_capacity = max_data_rows_per_sheet

                take = min(remaining_capacity, n - start)
                sub = chunk.iloc[start : start + take]
                write_df_rows(ws, sub)

                rows_in_part += take
                start += take

        # save last part
        outputs.append(str(save_part(wb, part_idx)))

        if parts_created > 1:
            return ("SPLIT", f"Split into {parts_created} file(s)", parts_created, ";".join(outputs))
        return ("OK", "", 1, ";".join(outputs))

    else:
        return ("FAILED", f"Unknown mode: {mode}", 0, "")


def main() -> None:
    ap = argparse.ArgumentParser(description="Convert TSVs to XLSX, splitting large files automatically.")
    ap.add_argument("folder", nargs="?", default=".", help="Folder containing .tsv files")
    ap.add_argument("--mode", choices=["sheets", "files"], default="sheets",
                    help="Split mode: one xlsx with many sheets OR many xlsx files")
    ap.add_argument("--chunksize", type=int, default=50_000, help="Rows per pandas chunk (speed/memory tradeoff)")
    args = ap.parse_args()

    folder = Path(args.folder).expanduser().resolve()
    if not folder.is_dir():
        raise SystemExit(f"Not a folder: {folder}")

    tsv_files = sorted(folder.glob("*.tsv"))
    if not tsv_files:
        print(f"No .tsv files found in: {folder}")
        return

    out_dir = folder / "xlsx"
    out_dir.mkdir(exist_ok=True)

    report_path = folder / "conversion_report.csv"
    with report_path.open("w", newline="", encoding="utf-8") as rf:
        writer = csv.writer(rf)
        writer.writerow(["file", "status", "reason", "parts", "outputs"])

        for tsv in tsv_files:
            try:
                status, reason, parts, outputs = convert_tsv_split(
                    tsv, out_dir=out_dir, mode=args.mode, chunksize=args.chunksize
                )
                writer.writerow([tsv.name, status, reason, parts, outputs])
                icon = "✅" if status == "OK" else ("🧩" if status == "SPLIT" else ("⚠️" if status == "SKIPPED" else "❌"))
                print(f"{icon} {tsv.name}: {status}" + (f" ({reason})" if reason else ""))
            except Exception as e:
                writer.writerow([tsv.name, "FAILED", repr(e), 0, ""])
                print(f"❌ {tsv.name}: FAILED ({e})")

    print(f"\nOutput folder: {out_dir}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()




pip install pandas openpyxl
python tsv_to_xlsx_split.py "D:\your\folder" --mode sheets
# or:
python tsv_to_xlsx_split.py "D:\your\folder" --mode files
