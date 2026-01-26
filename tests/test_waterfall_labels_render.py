from __future__ import annotations

import io
import zipfile
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook

from app import (
    _build_cell_range_formula,
    _find_header_column,
    _range_boundaries_from_formula,
    _worksheet_and_range_from_formula,
)
from test_utils import (
    build_deck_bytes,
    build_sample_dataframe,
    build_test_template,
    write_excel,
)


def _series_cache_points(root: ET.Element) -> list[tuple[int, int, int, int]]:
    ns = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
    points = []
    for ser in root.findall(".//c:ser", ns):
        cat_cache = ser.find("c:cat/c:strRef/c:strCache", ns)
        if cat_cache is None:
            cat_cache = ser.find("c:cat/c:numRef/c:numCache", ns)
        num_cache = ser.find("c:val/c:numRef/c:numCache", ns)
        cat_pt_count = 0
        cat_pts = 0
        if cat_cache is not None:
            pt_count = cat_cache.find("c:ptCount", ns)
            if pt_count is not None:
                cat_pt_count = int(pt_count.attrib.get("val", "0"))
            cat_pts = len(cat_cache.findall("c:pt", ns))
        num_pt_count = 0
        num_pts = 0
        if num_cache is not None:
            pt_count = num_cache.find("c:ptCount", ns)
            if pt_count is not None:
                num_pt_count = int(pt_count.attrib.get("val", "0"))
            num_pts = len(num_cache.findall("c:pt", ns))
        points.append((cat_pt_count, cat_pts, num_pt_count, num_pts))
    return points


def test_waterfall_labels_render_without_edit_data(tmp_path) -> None:
    template_path = tmp_path / "template.pptx"
    build_test_template(template_path, waterfall_slide_count=1)

    df = build_sample_dataframe(["Alpha"], include_brand=False)
    excel_path = tmp_path / "input.xlsx"
    write_excel(df, excel_path)
    df_from_excel = pd.read_excel(excel_path)
    pptx_bytes = build_deck_bytes(template_path, df_from_excel, waterfall_targets=["Alpha"])

    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zf:
        chart_files = [
            name
            for name in zf.namelist()
            if name.startswith("ppt/charts/chart") and name.endswith(".xml")
        ]
        assert chart_files
        chart_xml = zf.read(chart_files[0])

    root = ET.fromstring(chart_xml)
    ns = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
    d_lbls = root.findall(".//c:dLbls", ns)
    assert d_lbls

    label_flags = [
        lbl.find("c:showVal", ns) is not None or lbl.find("c:showCatName", ns) is not None
        for lbl in d_lbls
    ]
    assert any(label_flags)

    cache_points = _series_cache_points(root)
    assert cache_points
    for cat_pt_count, cat_pts, num_pt_count, num_pts in cache_points:
        assert cat_pt_count > 0
        assert cat_pts > 0
        assert num_pt_count > 0
        assert num_pts > 0


def _insert_value_from_cells_labels(template_path) -> None:
    template_bytes = template_path.read_bytes()
    with zipfile.ZipFile(io.BytesIO(template_bytes)) as zf:
        chart_files = [
            name
            for name in zf.namelist()
            if name.startswith("ppt/charts/chart") and name.endswith(".xml")
        ]
        assert chart_files
        chart_name = chart_files[0]
        chart_xml = zf.read(chart_name)
        embedding_files = [
            name
            for name in zf.namelist()
            if name.startswith("ppt/embeddings/") and name.endswith(".xlsx")
        ]
        assert embedding_files
        embedding_name = embedding_files[0]
        workbook_bytes = zf.read(embedding_name)

    ns_c = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    ns_c15 = "http://schemas.microsoft.com/office/drawing/2012/chart"
    ET.register_namespace("c", ns_c)
    ET.register_namespace("c15", ns_c15)
    root = ET.fromstring(chart_xml)
    root.set("xmlns:c15", ns_c15)
    ns = {"c": ns_c}
    series_nodes = root.findall(".//c:ser", ns)
    assert len(series_nodes) >= 2
    series_node = series_nodes[1]
    d_lbls = series_node.find("c:dLbls", ns)
    if d_lbls is None:
        d_lbls = ET.SubElement(series_node, f"{{{ns_c}}}dLbls")
    tx = ET.SubElement(d_lbls, f"{{{ns_c}}}tx")
    str_ref = ET.SubElement(tx, f"{{{ns_c}}}strRef")
    f_node = ET.SubElement(str_ref, f"{{{ns_c}}}f")
    f_node.text = "Sheet1!$H$2:$H$4"
    str_cache = ET.SubElement(str_ref, f"{{{ns_c}}}strCache")
    pt_count = ET.SubElement(str_cache, f"{{{ns_c}}}ptCount")
    pt_count.set("val", "3")
    for idx, value in enumerate(["old-1", "old-2", "old-3"]):
        pt = ET.SubElement(str_cache, f"{{{ns_c}}}pt", idx=str(idx))
        v = ET.SubElement(pt, f"{{{ns_c}}}v")
        v.text = value

    c15_range = ET.SubElement(d_lbls, f"{{{ns_c15}}}datalabelsRange")
    c15_f = ET.SubElement(c15_range, f"{{{ns_c15}}}f")
    c15_f.text = "Sheet1!$H$2:$H$4"
    c15_cache = ET.SubElement(c15_range, f"{{{ns_c15}}}dlblRangeCache")
    c15_pt_count = ET.SubElement(c15_cache, f"{{{ns_c15}}}ptCount")
    c15_pt_count.set("val", "3")
    for idx, value in enumerate(["old-1", "old-2", "old-3"]):
        pt = ET.SubElement(c15_cache, f"{{{ns_c15}}}pt", idx=str(idx))
        v = ET.SubElement(pt, f"{{{ns_c15}}}v")
        v.text = value

    updated_chart_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    workbook = load_workbook(io.BytesIO(workbook_bytes))
    ws = workbook.active
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value="labs-Positives")
    updated_workbook = io.BytesIO()
    workbook.save(updated_workbook)

    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(template_bytes)) as zf:
        with zipfile.ZipFile(output, "w") as zf_out:
            for info in zf.infolist():
                data = zf.read(info.filename)
                if info.filename == chart_name:
                    data = updated_chart_xml
                elif info.filename == embedding_name:
                    data = updated_workbook.getvalue()
                zf_out.writestr(info, data)
    template_path.write_bytes(output.getvalue())


def test_waterfall_c15_labels_render_without_edit_data(tmp_path) -> None:
    template_path = tmp_path / "template.pptx"
    build_test_template(template_path, waterfall_slide_count=1)
    _insert_value_from_cells_labels(template_path)

    df = build_sample_dataframe(["Alpha"], include_brand=False)
    excel_path = tmp_path / "input.xlsx"
    write_excel(df, excel_path)
    df_from_excel = pd.read_excel(excel_path)
    pptx_bytes = build_deck_bytes(template_path, df_from_excel, waterfall_targets=["Alpha"])

    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zf:
        chart_files = [
            name
            for name in zf.namelist()
            if name.startswith("ppt/charts/chart") and name.endswith(".xml")
        ]
        assert chart_files
        chart_xml = zf.read(chart_files[0])
        embedding_files = [
            name
            for name in zf.namelist()
            if name.startswith("ppt/embeddings/") and name.endswith(".xlsx")
        ]
        assert embedding_files
        workbook_bytes = zf.read(embedding_files[0])

    ns = {
        "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
        "c15": "http://schemas.microsoft.com/office/drawing/2012/chart",
    }
    root = ET.fromstring(chart_xml)
    series_node = root.findall(".//c:ser", ns)[1]
    c15_range = series_node.find(".//c15:datalabelsRange", ns)
    assert c15_range is not None
    c15_formula = c15_range.find("c15:f", ns)
    assert c15_formula is not None

    value_formula = series_node.find("c:val/c:numRef/c:f", ns)
    assert value_formula is not None and value_formula.text
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    ws, _, sheet_name = _worksheet_and_range_from_formula(workbook, value_formula.text)
    bounds = _range_boundaries_from_formula(value_formula.text)
    assert bounds is not None
    _, min_row, _, max_row = bounds
    labs_col = _find_header_column(ws, ["labs-Positives"])
    assert labs_col is not None
    expected_formula = _build_cell_range_formula(sheet_name, labs_col, min_row, max_row)

    assert c15_formula.text == expected_formula

    label_values = [
        ws.cell(row=row_idx, column=labs_col).value
        for row_idx in range(min_row, max_row + 1)
    ]
    expected_labels = ["" if value is None else str(value) for value in label_values]
    cache = c15_range.find("c15:dlblRangeCache", ns)
    assert cache is not None
    pt_count = cache.find("c15:ptCount", ns)
    assert pt_count is not None
    assert int(pt_count.attrib.get("val", "0")) == len(expected_labels)
    points = cache.findall("c15:pt", ns)
    assert len(points) == len(expected_labels)
    values = [pt.find("c15:v", ns).text if pt.find("c15:v", ns) is not None else "" for pt in points]
    assert values[0] == expected_labels[0]
    assert values[-1] == expected_labels[-1]
